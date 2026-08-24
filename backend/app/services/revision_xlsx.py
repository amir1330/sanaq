"""Build Excel workbook for a stock revision."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import StockRevision
from app.services.revisions import line_difference, line_value, revision_summary

INK = "1B1811"
LINE = "DDD5C1"
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor=INK)
title_font = Font(name="Calibri", bold=True, size=14, color=INK)
meta_font = Font(name="Calibri", size=11, color="5B5346")
thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)


def _num(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return 0.0


def _when(value: datetime | None) -> str:
    if value is None:
        return "—"
    return value.astimezone().strftime("%d.%m.%Y %H:%M")


def build_revision_xlsx(
    revision: StockRevision,
    *,
    shop_name: str,
    created_by_name: str | None,
    posted_by_name: str | None,
) -> bytes:
    summary = revision_summary(revision, hide_cost=False)
    wb = Workbook()
    cover = wb.active
    cover.title = "Сводка"
    cover["A1"] = f"Ревизия №{revision.id} · {shop_name}"
    cover["A1"].font = title_font
    cover.merge_cells("A1:B1")
    rows = [
        ("Статус", {"draft": "черновик", "posted": "проведена", "cancelled": "отменена"}.get(revision.status.value, revision.status.value)),
        ("Снимок системы", _when(revision.created_at)),
        ("Проведена / отменена", _when(revision.posted_at or revision.cancelled_at)),
        ("Открыл", created_by_name or "—"),
        ("Закрыл", posted_by_name or "—"),
        ("Комментарий", revision.comment or "—"),
        ("Позиций", summary["line_count"]),
        ("Посчитано", summary["counted_count"]),
        ("Недостача, шт. поз.", summary["shortage_count"]),
        ("Излишек, шт. поз.", summary["surplus_count"]),
        ("Расхождение, ₸", _num(summary["difference_value"])),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        cover[f"A{i}"] = label
        cover[f"A{i}"].font = meta_font
        cover[f"B{i}"] = value
    cover.column_dimensions["A"].width = 28
    cover.column_dimensions["B"].width = 36

    sheet = wb.create_sheet("Строки")
    headers = ["Позиция", "Ед.", "Система", "Факт", "Δ", "Себест./ед.", "Сумма Δ, ₸", "Заметка"]
    for col, title in enumerate(headers, start=1):
        cell = sheet.cell(1, col, title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin

    ordered = sorted(revision.lines, key=lambda row: (row.stock_item_name, row.id))
    for r_idx, line in enumerate(ordered, start=2):
        diff = line.difference_quantity
        if line.counted_quantity is not None and diff is None:
            diff = line_difference(line.counted_quantity, line.expected_quantity)
        value = line_value(diff, line.cost_per_base_unit)
        values = [
            line.stock_item_name,
            line.base_unit,
            _num(line.expected_quantity),
            None if line.counted_quantity is None else _num(line.counted_quantity),
            None if diff is None else _num(diff),
            _num(line.cost_per_base_unit),
            None if value is None else _num(value),
            line.comment or "",
        ]
        for c_idx, value in enumerate(values, start=1):
            cell = sheet.cell(r_idx, c_idx, value)
            cell.border = thin
            if c_idx >= 3 and c_idx <= 7 and isinstance(value, (int, float)):
                cell.number_format = "#,##0.###" if c_idx < 6 else "#,##0.00"

    widths = [28, 8, 12, 12, 12, 14, 14, 28]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
