from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings


INK = "1B1811"
LINE = "DDD5C1"
GOLD = "B08A2E"

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor=INK)
title_font = Font(name="Calibri", bold=True, size=14, color=INK)
meta_font = Font(name="Calibri", size=11, color="5B5346")
money_font = Font(name="Calibri", size=11)
thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
header_align = Alignment(horizontal="left", vertical="center")


def _num(value: object) -> float | int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float):
        return value
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def _when(value: datetime | date | str | None) -> datetime | date | str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    return str(value)[:16]


def _style_header(ws: Worksheet, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin


def _write_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[object]],
    money_cols: set[int] | None = None,
) -> None:
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"

    money_cols = money_cols or set()
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = thin
            cell.font = money_font
            if c_idx in money_cols and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            if isinstance(value, datetime):
                cell.number_format = "DD.MM.YYYY HH:MM"
            elif isinstance(value, date):
                cell.number_format = "DD.MM.YYYY"

    widths = [len(h) for h in headers]
    for row in rows:
        for i, value in enumerate(row):
            widths[i] = max(widths[i], min(40, len(str(value))))
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, width + 3)

    ws.row_dimensions[1].height = 22


def build_report_xlsx(
    *,
    shop_name: str,
    from_date: date,
    to_date: date,
    summary: dict[str, object],
    daily: list[dict[str, object]],
    products: list[dict[str, object]],
    sellers: list[dict[str, object]],
    expenses: list[dict[str, object]],
    sales: list[dict[str, object]],
) -> bytes:
    wb = Workbook()
    cover = wb.active
    cover.title = "Сводка"
    cover.sheet_view.showGridLines = False
    cover["A1"] = settings.app_name
    cover["A1"].font = title_font
    cover["A2"] = shop_name
    cover["A2"].font = meta_font
    cover["A3"] = f"{from_date.isoformat()} — {to_date.isoformat()}"
    cover["A3"].font = meta_font

    cover.append([])
    cover.append(["Показатель", "Сумма"])
    _style_header(cover, 5, 2)
    metrics = [
        ("Выручка", _num(summary.get("revenue"))),
        ("Наличными", _num(summary.get("cash_revenue"))),
        ("Безналично", _num(summary.get("card_revenue"))),
        ("Себестоимость", _num(summary.get("cost"))),
        ("Прибыль", _num(summary.get("profit"))),
        ("Расходы", _num(summary.get("expenses"))),
        ("Недостачи по ревизиям", _num(summary.get("revision_shortage"))),
        ("Чистыми", _num(summary.get("net_profit"))),
        ("Чеков", int(summary.get("sales_count") or 0)),
    ]
    for i, (label, value) in enumerate(metrics, start=6):
        cover.cell(i, 1, label).border = thin
        cell = cover.cell(i, 2, value)
        cell.border = thin
        if label != "Чеков":
            cell.number_format = '#,##0.00'
            if label in {"Прибыль", "Чистыми"}:
                cell.font = Font(name="Calibri", bold=True, color=GOLD)
    cover.column_dimensions["A"].width = 26
    cover.column_dimensions["B"].width = 16
    cover.freeze_panes = "A6"

    _write_sheet(
        wb,
        "По дням",
        ["День", "Наличными", "Безналично", "Выручка", "Себестоимость", "Прибыль", "Чеки"],
        [
            [
                _when(row.get("day")),
                _num(row.get("cash_revenue")),
                _num(row.get("card_revenue")),
                _num(row.get("revenue")),
                _num(row.get("cost")),
                _num(row.get("profit")),
                int(row.get("sales_count") or 0),
            ]
            for row in daily
        ],
        money_cols={2, 3, 4, 5, 6},
    )
    _write_sheet(
        wb,
        "Товары",
        ["Товар", "Штук", "Выручка", "Прибыль"],
        [
            [
                row.get("name") or "",
                int(row.get("quantity") or 0),
                _num(row.get("revenue")),
                _num(row.get("profit")),
            ]
            for row in products
        ],
        money_cols={3, 4},
    )
    _write_sheet(
        wb,
        "Продавцы",
        ["Имя", "Наличными", "Безналично", "Выручка", "Чеки"],
        [
            [
                row.get("barista_name") or "",
                _num(row.get("cash_revenue")),
                _num(row.get("card_revenue")),
                _num(row.get("revenue")),
                int(row.get("sales_count") or 0),
            ]
            for row in sellers
        ],
        money_cols={2, 3, 4},
    )
    _write_sheet(
        wb,
        "Расходы",
        ["Дата", "Категория", "Сумма", "Комментарий"],
        [
            [
                _when(row.get("created_at")),
                row.get("category") or "",
                _num(row.get("amount")),
                row.get("comment") or "",
            ]
            for row in expenses
        ],
        money_cols={3},
    )
    _write_sheet(
        wb,
        "Чеки",
        ["Дата", "Номер", "Продавец", "Оплата", "Сумма", "Возврат", "Фискализация"],
        [
            [
                _when(row.get("created_at")),
                row.get("id"),
                row.get("barista_name") or "",
                "Наличными" if row.get("payment_type") == "cash" else "Безналично",
                _num(row.get("total_amount")),
                "да" if row.get("is_refunded") else "",
                row.get("fiscal_status") or "",
            ]
            for row in sales
        ],
        money_cols={5},
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
