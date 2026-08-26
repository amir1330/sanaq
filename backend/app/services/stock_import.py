"""Excel import for stock items."""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.stock import StockImportPreviewOut, StockImportPreviewRow, StockImportRowIn

REQUIRED_FIELDS = (
    "name",
    "base_unit",
    "purchase_unit",
    "purchase_to_base",
    "quantity",
    "cost_per_base_unit",
)
OPTIONAL_FIELDS = ("min_quantity", "is_ingredient")
ALL_FIELDS = REQUIRED_FIELDS + OPTIONAL_FIELDS

# Human-readable headers per locale (also accepted when re-uploading).
LABELS: dict[str, dict[str, str]] = {
    "ru": {
        "name": "Название",
        "base_unit": "Ед. остатка",
        "purchase_unit": "Ед. закупки",
        "purchase_to_base": "Сколько ед. остатка в 1 ед. закупки",
        "quantity": "Начальный остаток (ед. закупки)",
        "cost_per_base_unit": "Себестоимость за 1 ед. остатка",
        "min_quantity": "Минимум (ед. остатка)",
        "is_ingredient": "Только ингредиент (да/нет)",
    },
    "en": {
        "name": "Name",
        "base_unit": "Stock unit",
        "purchase_unit": "Purchase unit",
        "purchase_to_base": "Stock units in 1 purchase unit",
        "quantity": "Opening qty (purchase units)",
        "cost_per_base_unit": "Cost per stock unit",
        "min_quantity": "Min stock (stock units)",
        "is_ingredient": "Ingredient only (yes/no)",
    },
    "kk": {
        "name": "Атауы",
        "base_unit": "Қалдық бірлігі",
        "purchase_unit": "Сатып алу бірлігі",
        "purchase_to_base": "1 сатып алуда қанша қалдық",
        "quantity": "Бастапқы қалдық (сатып алу)",
        "cost_per_base_unit": "1 қалдық бірлігінің өзіндік құны",
        "min_quantity": "Минимум (қалдық)",
        "is_ingredient": "Тек ингредиент (иә/жоқ)",
    },
}

GUIDE: dict[str, dict[str, Any]] = {
    "ru": {
        "sheet": "Склад",
        "guide_sheet": "Инструкция",
        "title": "Как заполнять",
        "lines": [
            "Жёлтые колонки обязательны. Белые — по желанию.",
            "Название — как позиция называется на складе.",
            "Ед. остатка — в чём считаем остаток (мл, г, шт).",
            "Ед. закупки — как покупаете (л, кг, пачка, шт).",
            "Сколько ед. остатка в 1 ед. закупки — например, 1 л = 1000 мл → 1000.",
            "Начальный остаток — сколько сейчас на складе в единицах закупки.",
            "Себестоимость за 1 ед. остатка — цена одной мл/г/шт.",
            "Минимум — порог «мало» в ед. остатка (необязательно).",
            "Только ингредиент — да/нет: не предлагать «сделать товаром» на кассе.",
            "Строки 2–3 на листе «Склад» — примеры: удалите или замените своими данными.",
        ],
    },
    "en": {
        "sheet": "Stock",
        "guide_sheet": "Guide",
        "title": "How to fill this in",
        "lines": [
            "Yellow columns are required. White ones are optional.",
            "Name — stock item name.",
            "Stock unit — unit for on-hand balance (ml, g, pcs).",
            "Purchase unit — how you buy it (L, kg, pack, pcs).",
            "Stock units in 1 purchase unit — e.g. 1 L = 1000 ml → 1000.",
            "Opening qty — current stock in purchase units.",
            "Cost per stock unit — cost of one ml/g/pc.",
            "Min stock — low-stock threshold in stock units (optional).",
            "Ingredient only — yes/no: hide from “sell on till”.",
            "Rows 2–3 on the Stock sheet are examples — delete or replace them.",
        ],
    },
    "kk": {
        "sheet": "Қойма",
        "guide_sheet": "Нұсқау",
        "title": "Қалай толтыру керек",
        "lines": [
            "Сары бағандар міндетті. Ақ бағандар — қалауыңызша.",
            "Атауы — қоймадағы позиция атауы.",
            "Қалдық бірлігі — қалдықты неде санаймыз (мл, г, дана).",
            "Сатып алу бірлігі — қалай сатып аласыз (л, кг, пакет, дана).",
            "1 сатып алуда қанша қалдық — мыс. 1 л = 1000 мл → 1000.",
            "Бастапқы қалдық — қазір қоймада сатып алу бірлігінде.",
            "1 қалдық бірлігінің өзіндік құны — бір мл/г/дана бағасы.",
            "Минимум — «аз» шегі қалдық бірлігінде (міндетті емес).",
            "Тек ингредиент — иә/жоқ: кассаға «тауарға айналдыру» ұсынбау.",
            "«Қойма» парағындағы 2–3 жол — мысал: жойыңыз немесе өз деректеріңізбен ауыстырыңыз.",
        ],
    },
}

YES_NO: dict[str, tuple[str, str]] = {
    "ru": ("да", "нет"),
    "en": ("yes", "no"),
    "kk": ("иә", "жоқ"),
}

YELLOW = PatternFill("solid", fgColor="FFE599")
HEADER_FONT = Font(bold=True)


def _norm_header(value: str) -> str:
    return " ".join(value.strip().lower().split())


def _header_aliases() -> dict[str, str]:
    """Map any known header text → field key."""
    aliases: dict[str, str] = {}
    for field in ALL_FIELDS:
        aliases[_norm_header(field)] = field
        for labels in LABELS.values():
            aliases[_norm_header(labels[field])] = field
    return aliases


HEADER_ALIASES = _header_aliases()


def _resolve_locale(lang: str | None) -> str:
    if lang and lang.lower()[:2] in LABELS:
        return lang.lower()[:2]
    return "ru"


def build_stock_import_template(lang: str | None = "ru") -> bytes:
    locale = _resolve_locale(lang)
    labels = LABELS[locale]
    guide = GUIDE[locale]
    yes, no = YES_NO[locale]

    wb = Workbook()
    ws = wb.active
    ws.title = guide["sheet"]
    for col, field in enumerate(ALL_FIELDS, start=1):
        cell = ws.cell(1, col, labels[field])
        cell.font = HEADER_FONT
        if field in REQUIRED_FIELDS:
            cell.fill = YELLOW
        width = max(14, min(42, len(labels[field]) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width

    example = {
        "name": "Молоко 3.2%" if locale != "en" else "Milk 3.2%",
        "base_unit": "мл" if locale != "en" else "ml",
        "purchase_unit": "л" if locale != "en" else "L",
        "purchase_to_base": 1000,
        "quantity": 12,
        "cost_per_base_unit": 0.45,
        "min_quantity": 2000,
        "is_ingredient": yes,
    }
    example2 = {
        "name": "Вода 0.5л" if locale != "en" else "Water 0.5L",
        "base_unit": "шт" if locale != "en" else "pcs",
        "purchase_unit": "шт" if locale != "en" else "pcs",
        "purchase_to_base": 1,
        "quantity": 48,
        "cost_per_base_unit": 80,
        "min_quantity": 12,
        "is_ingredient": no,
    }
    for col, field in enumerate(ALL_FIELDS, start=1):
        ws.cell(2, col, example[field])
        ws.cell(3, col, example2[field])

    guide_ws = wb.create_sheet(guide["guide_sheet"])
    guide_ws["A1"] = guide["title"]
    guide_ws["A1"].font = HEADER_FONT
    for i, line in enumerate(guide["lines"], start=3):
        guide_ws[f"A{i}"] = line
    guide_ws.column_dimensions["A"].width = 96

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_bool(raw: str) -> bool:
    v = raw.strip().lower()
    return v in {"1", "true", "yes", "y", "да", "д", "иә", "iae", "иа"}


def _parse_decimal(raw: str, *, field: str, errors: list[str], label: str) -> Decimal | None:
    if not raw:
        errors.append(f"{label}: пусто")
        return None
    try:
        return Decimal(raw.replace(",", ".").replace(" ", ""))
    except (InvalidOperation, ValueError):
        errors.append(f"{label}: не число")
        return None


def _map_headers(header_row: list[str]) -> dict[str, int]:
    index: dict[str, int] = {}
    for i, raw in enumerate(header_row):
        key = HEADER_ALIASES.get(_norm_header(raw))
        if key and key not in index:
            index[key] = i
    return index


def parse_stock_import_xlsx(content: bytes) -> StockImportPreviewOut:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    header_row = [_cell_str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    index = _map_headers(header_row)
    labels = LABELS["ru"]
    missing = [labels[h] for h in REQUIRED_FIELDS if h not in index]
    if missing:
        return StockImportPreviewOut(
            rows=[
                StockImportPreviewRow(
                    row=1,
                    ok=False,
                    errors=[f"Нет колонок: {', '.join(missing)}"],
                )
            ],
            ok_count=0,
            error_count=1,
        )

    rows: list[StockImportPreviewRow] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [_cell_str(c.value) for c in row]
        if not any(values):
            continue
        errors: list[str] = []

        def cell(field: str) -> str:
            i = index.get(field)
            if i is None or i >= len(values):
                return ""
            return values[i]

        name = cell("name")
        if not name:
            errors.append(f"{labels['name']}: пусто")

        base_unit = cell("base_unit")
        purchase_unit = cell("purchase_unit")
        if not base_unit:
            errors.append(f"{labels['base_unit']}: пусто")
        if not purchase_unit:
            errors.append(f"{labels['purchase_unit']}: пусто")

        factor = _parse_decimal(
            cell("purchase_to_base"),
            field="purchase_to_base",
            errors=errors,
            label=labels["purchase_to_base"],
        )
        qty = _parse_decimal(
            cell("quantity"),
            field="quantity",
            errors=errors,
            label=labels["quantity"],
        )
        cost = _parse_decimal(
            cell("cost_per_base_unit"),
            field="cost_per_base_unit",
            errors=errors,
            label=labels["cost_per_base_unit"],
        )
        min_raw = cell("min_quantity")
        min_qty = Decimal("0")
        if min_raw:
            parsed_min = _parse_decimal(
                min_raw,
                field="min_quantity",
                errors=errors,
                label=labels["min_quantity"],
            )
            if parsed_min is not None:
                min_qty = parsed_min

        is_ing = _parse_bool(cell("is_ingredient")) if "is_ingredient" in index else False

        if factor is not None and factor <= 0:
            errors.append(f"{labels['purchase_to_base']}: должно быть > 0")
        if qty is not None and qty < 0:
            errors.append(f"{labels['quantity']}: не может быть < 0")
        if cost is not None and cost < 0:
            errors.append(f"{labels['cost_per_base_unit']}: не может быть < 0")

        data = None
        ok = not errors and factor is not None and qty is not None and cost is not None and bool(name)
        if ok:
            data = StockImportRowIn(
                name=name,
                base_unit=base_unit,
                purchase_unit=purchase_unit,
                purchase_to_base=factor,
                quantity=qty,
                cost_per_base_unit=cost,
                min_quantity=min_qty,
                is_ingredient=is_ing,
            )
        rows.append(StockImportPreviewRow(row=r_idx, ok=ok, errors=errors, data=data))

    ok_count = sum(1 for r in rows if r.ok)
    return StockImportPreviewOut(
        rows=rows,
        ok_count=ok_count,
        error_count=len(rows) - ok_count,
    )
