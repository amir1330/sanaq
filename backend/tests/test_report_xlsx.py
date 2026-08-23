from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.services.reports import build_report_xlsx


def test_report_xlsx_has_sheets_and_numbers():
    raw = build_report_xlsx(
        shop_name="Corner",
        from_date=date(2026, 8, 1),
        to_date=date(2026, 8, 23),
        summary={
            "revenue": Decimal("120.00"),
            "cash_revenue": Decimal("80"),
            "card_revenue": Decimal("40"),
            "cost": Decimal("30"),
            "profit": Decimal("90"),
            "expenses": Decimal("10"),
            "net_profit": Decimal("80"),
            "sales_count": 3,
        },
        daily=[
            {
                "day": datetime(2026, 8, 23, 0, 0, tzinfo=timezone.utc),
                "cash_revenue": Decimal("80"),
                "card_revenue": Decimal("40"),
                "revenue": Decimal("120"),
                "cost": Decimal("30"),
                "profit": Decimal("90"),
                "sales_count": 3,
            }
        ],
        products=[{"name": "Латте", "quantity": 2, "revenue": Decimal("10"), "profit": Decimal("4")}],
        sellers=[
            {
                "barista_name": "Amina",
                "cash_revenue": Decimal("80"),
                "card_revenue": Decimal("0"),
                "revenue": Decimal("80"),
                "sales_count": 2,
            }
        ],
        expenses=[
            {
                "created_at": datetime(2026, 8, 10, 12, 0, tzinfo=timezone.utc),
                "category": "Аренда",
                "amount": Decimal("10"),
                "comment": "август",
            }
        ],
        sales=[
            {
                "created_at": datetime(2026, 8, 23, 9, 15, tzinfo=timezone.utc),
                "id": 11,
                "barista_name": "Amina",
                "payment_type": "cash",
                "total_amount": Decimal("4.50"),
                "is_refunded": False,
            }
        ],
    )
    wb = load_workbook(BytesIO(raw))
    assert wb.sheetnames == ["Сводка", "По дням", "Товары", "Продавцы", "Расходы", "Чеки"]
    cover = wb["Сводка"]
    assert cover["A2"].value == "Corner"
    assert cover["B6"].value == 120
    assert cover["B6"].number_format == "#,##0.00"
    assert wb["Товары"]["A2"].value == "Латте"
    assert wb["Продавцы"]["A2"].value == "Amina"
    assert wb["Расходы"]["B2"].value == "Аренда"
    assert wb["Чеки"]["D2"].value == "Наличными"
    assert wb["Чеки"]["E2"].value == 4.5
